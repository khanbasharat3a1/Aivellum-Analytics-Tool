"""
Financial Management Module for Aivellum
Handles Income, Expenses, Salaries, and Planned Activities
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class FinancialManager:
    def __init__(self, file_path='Aivellum_Financials_OCT.xlsx'):
        self.file_path = file_path
        self.sheets = {
            'Income': ['Income Source', 'Amount', 'Date', 'Category', 'Notes'],
            'Expenses': ['Description', 'Amount', 'Date', 'Category', 'Type', 'Notes'],
            'Planned': ['Activity', 'Estimated Cost', 'Priority', 'Status', 'Target Date', 'Notes']
        }
        self.expense_categories = {
            'Salaries': ['Basharat', 'Abdaal', 'Saleem'],
            'Outsourcing': ['Editor'],
            'Tools': ['Cursor Pro', 'Software'],
            'Business': ['Domain', 'Hosting', 'API Credits'],
            'Marketing': ['Ads', 'Promotion']
        }
        self.init_file()
    
    def init_file(self):
        """Initialize Excel file with sample data if it doesn't exist"""
        if not os.path.exists(self.file_path):
            try:
                self.create_sample_data()
            except Exception as e:
                print(f"Warning: Could not create Excel file: {e}")
                # Create empty DataFrames as fallback
                self._create_empty_data()
    
    def _create_empty_data(self):
        """Create empty data structure as fallback"""
        try:
            import pandas as pd
            
            # Create empty DataFrames with proper structure
            income_df = pd.DataFrame(columns=['Income Source', 'Amount', 'Date', 'Category', 'Notes'])
            expense_df = pd.DataFrame(columns=['Description', 'Amount', 'Date', 'Category', 'Type', 'Notes'])
            planned_df = pd.DataFrame(columns=['Activity', 'Estimated Cost', 'Priority', 'Status', 'Target Date', 'Notes'])
            
            # Try to save as Excel, fallback to CSV if needed
            try:
                with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                    income_df.to_excel(writer, sheet_name='Income', index=False)
                    expense_df.to_excel(writer, sheet_name='Expenses', index=False)
                    planned_df.to_excel(writer, sheet_name='Planned', index=False)
            except ImportError:
                # Fallback: create CSV files
                income_df.to_csv(self.file_path.replace('.xlsx', '_income.csv'), index=False)
                expense_df.to_csv(self.file_path.replace('.xlsx', '_expenses.csv'), index=False)
                planned_df.to_csv(self.file_path.replace('.xlsx', '_planned.csv'), index=False)
                print("Created CSV files instead of Excel")
        except Exception as e:
            print(f"Could not create data files: {e}")
    
    def create_sample_data(self):
        """Create Excel file with sample financial data"""
        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            
            # Income sheet
            ws_income = wb.active
            ws_income.title = "Income"
            income_data = [
                ['Income Source', 'Amount', 'Date', 'Category', 'Notes'],
                ['Gumroad', 2070, '2025-11-03', 'Digital Products', 'Course sales'],
                ['Gumroad', 2080, '2025-11-10', 'Digital Products', 'Template sales'],
                ['Play Console', 9060, '2025-11-17', 'App Revenue', 'Monthly payout'],
                ['Runnable', 2925, '2025-11-18', 'Services', 'Development work'],
                ['Runnable', 2948, '2025-11-26', 'Services', 'Consulting']
            ]
            for row in income_data:
                ws_income.append(row)
            
            # Expenses sheet
            ws_expenses = wb.create_sheet("Expenses")
            expense_data = [
                ['Description', 'Amount', 'Date', 'Category', 'Type', 'Notes'],
                ['Cursor Pro', 2088.64, '2025-11-07', 'Tools', 'Software', 'Annual subscription'],
                ['Editor Payment', 500, '2025-11-11', 'Outsourcing', 'Service', 'Content editing'],
                ['Editor Payment', 500, '2025-11-25', 'Outsourcing', 'Service', 'Content editing'],
                ['Basharat Salary', 4000, '2025-11-30', 'Salaries', 'Employee', 'Monthly salary'],
                ['Abdaal Salary', 4000, '2025-11-30', 'Salaries', 'Employee', 'Monthly salary'],
                ['Editor Payment', 500, '2025-11-30', 'Outsourcing', 'Service', 'Content editing'],
                ['Saleem Salary', 4000, '2025-11-30', 'Salaries', 'Employee', 'Monthly salary']
            ]
            for row in expense_data:
                ws_expenses.append(row)
            
            # Planned sheet
            ws_planned = wb.create_sheet("Planned")
            planned_data = [
                ['Activity', 'Estimated Cost', 'Priority', 'Status', 'Target Date', 'Notes'],
                ['Zoho Workplace (5 accounts)', 750, 'High', 'Pending', '2025-12-15', 'Email & collaboration'],
                ['Domain Renewal', 1000, 'High', 'Pending', '2025-12-31', 'Annual renewal'],
                ['Cursor Pro Renewal', 2000, 'Medium', 'Pending', '2026-11-07', 'Development tool'],
                ['Editors Monthly Salary', 5000, 'High', 'Recurring', '2025-12-01', 'Content team'],
                ['OpenAI API Credits', 750, 'Medium', 'Pending', '2025-12-10', 'AI services'],
                ['Marketing Ads', 1000, 'Low', 'Planned', '2025-12-20', 'Promotion campaign']
            ]
            for row in planned_data:
                ws_planned.append(row)
            
            wb.save(self.file_path)
            
        except ImportError:
            # Fallback: create CSV files with sample data
            import pandas as pd
            
            # Sample data
            income_data = {
                'Income Source': ['Gumroad', 'Gumroad', 'Play Console', 'Runnable', 'Runnable'],
                'Amount': [2070, 2080, 9060, 2925, 2948],
                'Date': ['2025-11-03', '2025-11-10', '2025-11-17', '2025-11-18', '2025-11-26'],
                'Category': ['Digital Products', 'Digital Products', 'App Revenue', 'Services', 'Services'],
                'Notes': ['Course sales', 'Template sales', 'Monthly payout', 'Development work', 'Consulting']
            }
            
            expense_data = {
                'Description': ['Cursor Pro', 'Editor Payment', 'Editor Payment', 'Basharat Salary', 'Abdaal Salary', 'Editor Payment', 'Saleem Salary'],
                'Amount': [2088.64, 500, 500, 4000, 4000, 500, 4000],
                'Date': ['2025-11-07', '2025-11-11', '2025-11-25', '2025-11-30', '2025-11-30', '2025-11-30', '2025-11-30'],
                'Category': ['Tools', 'Outsourcing', 'Outsourcing', 'Salaries', 'Salaries', 'Outsourcing', 'Salaries'],
                'Type': ['Software', 'Service', 'Service', 'Employee', 'Employee', 'Service', 'Employee'],
                'Notes': ['Annual subscription', 'Content editing', 'Content editing', 'Monthly salary', 'Monthly salary', 'Content editing', 'Monthly salary']
            }
            
            planned_data = {
                'Activity': ['Zoho Workplace (5 accounts)', 'Domain Renewal', 'Cursor Pro Renewal', 'Editors Monthly Salary', 'OpenAI API Credits', 'Marketing Ads'],
                'Estimated Cost': [750, 1000, 2000, 5000, 750, 1000],
                'Priority': ['High', 'High', 'Medium', 'High', 'Medium', 'Low'],
                'Status': ['Pending', 'Pending', 'Pending', 'Recurring', 'Pending', 'Planned'],
                'Target Date': ['2025-12-15', '2025-12-31', '2026-11-07', '2025-12-01', '2025-12-10', '2025-12-20'],
                'Notes': ['Email & collaboration', 'Annual renewal', 'Development tool', 'Content team', 'AI services', 'Promotion campaign']
            }
            
            # Create CSV files
            pd.DataFrame(income_data).to_csv(self.file_path.replace('.xlsx', '_income.csv'), index=False)
            pd.DataFrame(expense_data).to_csv(self.file_path.replace('.xlsx', '_expenses.csv'), index=False)
            pd.DataFrame(planned_data).to_csv(self.file_path.replace('.xlsx', '_planned.csv'), index=False)
            
            print("Created CSV files with sample financial data")
    
    def load_data(self):
        """Load all financial data"""
        try:
            import pandas as pd
            data = {}
            
            # Try to load Excel file first
            if os.path.exists(self.file_path):
                try:
                    for sheet_name in ['Income', 'Expenses', 'Planned']:
                        df = pd.read_excel(self.file_path, sheet_name=sheet_name)
                        if 'Date' in df.columns:
                            df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
                        elif 'Target Date' in df.columns:
                            df['Target Date'] = pd.to_datetime(df['Target Date'], errors='coerce')
                        data[sheet_name.lower()] = df
                    return data
                except Exception as e:
                    print(f"Error loading Excel: {e}")
            
            # Fallback: try CSV files
            csv_files = {
                'income': self.file_path.replace('.xlsx', '_income.csv'),
                'expenses': self.file_path.replace('.xlsx', '_expenses.csv'),
                'planned': self.file_path.replace('.xlsx', '_planned.csv')
            }
            
            for key, csv_file in csv_files.items():
                if os.path.exists(csv_file):
                    df = pd.read_csv(csv_file)
                    if 'Date' in df.columns:
                        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
                    elif 'Target Date' in df.columns:
                        df['Target Date'] = pd.to_datetime(df['Target Date'], errors='coerce')
                    data[key] = df
                else:
                    # Create empty DataFrame with proper columns
                    if key == 'income':
                        data[key] = pd.DataFrame(columns=['Income Source', 'Amount', 'Date', 'Category', 'Notes'])
                    elif key == 'expenses':
                        data[key] = pd.DataFrame(columns=['Description', 'Amount', 'Date', 'Category', 'Type', 'Notes'])
                    else:  # planned
                        data[key] = pd.DataFrame(columns=['Activity', 'Estimated Cost', 'Priority', 'Status', 'Target Date', 'Notes'])
            
            return data
            
        except Exception as e:
            print(f"Error loading financial data: {e}")
            # Return empty DataFrames with proper structure
            import pandas as pd
            return {
                'income': pd.DataFrame(columns=['Income Source', 'Amount', 'Date', 'Category', 'Notes']),
                'expenses': pd.DataFrame(columns=['Description', 'Amount', 'Date', 'Category', 'Type', 'Notes']),
                'planned': pd.DataFrame(columns=['Activity', 'Estimated Cost', 'Priority', 'Status', 'Target Date', 'Notes'])
            }
    
    def add_income(self, source, amount, date, category='Other', notes=''):
        """Add new income entry"""
        try:
            import pandas as pd
            
            # Try Excel first
            if os.path.exists(self.file_path):
                try:
                    df = pd.read_excel(self.file_path, sheet_name='Income')
                    new_entry = {
                        'Income Source': source,
                        'Amount': float(amount),
                        'Date': pd.to_datetime(date),
                        'Category': category,
                        'Notes': notes
                    }
                    df = pd.concat([df, pd.DataFrame([new_entry])], ignore_index=True)
                    
                    with pd.ExcelWriter(self.file_path, mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name='Income', index=False)
                    return True
                except Exception:
                    pass
            
            # Fallback to CSV
            csv_file = self.file_path.replace('.xlsx', '_income.csv')
            if os.path.exists(csv_file):
                df = pd.read_csv(csv_file)
            else:
                df = pd.DataFrame(columns=['Income Source', 'Amount', 'Date', 'Category', 'Notes'])
            
            new_entry = {
                'Income Source': source,
                'Amount': float(amount),
                'Date': date,
                'Category': category,
                'Notes': notes
            }
            df = pd.concat([df, pd.DataFrame([new_entry])], ignore_index=True)
            df.to_csv(csv_file, index=False)
            return True
            
        except Exception as e:
            print(f"Error adding income: {e}")
            return False
    
    def add_expense(self, description, amount, date, category, exp_type, notes=''):
        """Add new expense entry"""
        try:
            import pandas as pd
            
            # Try Excel first
            if os.path.exists(self.file_path):
                try:
                    df = pd.read_excel(self.file_path, sheet_name='Expenses')
                    new_entry = {
                        'Description': description,
                        'Amount': float(amount),
                        'Date': pd.to_datetime(date),
                        'Category': category,
                        'Type': exp_type,
                        'Notes': notes
                    }
                    df = pd.concat([df, pd.DataFrame([new_entry])], ignore_index=True)
                    
                    with pd.ExcelWriter(self.file_path, mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name='Expenses', index=False)
                    return True
                except Exception:
                    pass
            
            # Fallback to CSV
            csv_file = self.file_path.replace('.xlsx', '_expenses.csv')
            if os.path.exists(csv_file):
                df = pd.read_csv(csv_file)
            else:
                df = pd.DataFrame(columns=['Description', 'Amount', 'Date', 'Category', 'Type', 'Notes'])
            
            new_entry = {
                'Description': description,
                'Amount': float(amount),
                'Date': date,
                'Category': category,
                'Type': exp_type,
                'Notes': notes
            }
            df = pd.concat([df, pd.DataFrame([new_entry])], ignore_index=True)
            df.to_csv(csv_file, index=False)
            return True
            
        except Exception as e:
            print(f"Error adding expense: {e}")
            return False
    
    def add_planned_activity(self, activity, cost, priority, status, target_date, notes=''):
        """Add new planned activity"""
        try:
            import pandas as pd
            
            # Try Excel first
            if os.path.exists(self.file_path):
                try:
                    df = pd.read_excel(self.file_path, sheet_name='Planned')
                    new_entry = {
                        'Activity': activity,
                        'Estimated Cost': float(cost),
                        'Priority': priority,
                        'Status': status,
                        'Target Date': pd.to_datetime(target_date),
                        'Notes': notes
                    }
                    df = pd.concat([df, pd.DataFrame([new_entry])], ignore_index=True)
                    
                    with pd.ExcelWriter(self.file_path, mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name='Planned', index=False)
                    return True
                except Exception:
                    pass
            
            # Fallback to CSV
            csv_file = self.file_path.replace('.xlsx', '_planned.csv')
            if os.path.exists(csv_file):
                df = pd.read_csv(csv_file)
            else:
                df = pd.DataFrame(columns=['Activity', 'Estimated Cost', 'Priority', 'Status', 'Target Date', 'Notes'])
            
            new_entry = {
                'Activity': activity,
                'Estimated Cost': float(cost),
                'Priority': priority,
                'Status': status,
                'Target Date': target_date,
                'Notes': notes
            }
            df = pd.concat([df, pd.DataFrame([new_entry])], ignore_index=True)
            df.to_csv(csv_file, index=False)
            return True
            
        except Exception as e:
            print(f"Error adding planned activity: {e}")
            return False
    
    def get_financial_summary(self, start_date=None, end_date=None):
        """Get financial summary with cash flow analysis"""
        data = self.load_data()
        
        # Filter by date if provided
        if start_date and end_date:
            start_dt = pd.to_datetime(start_date)
            end_dt = pd.to_datetime(end_date)
            
            if not data['income'].empty and 'Date' in data['income'].columns:
                data['income'] = data['income'][
                    (data['income']['Date'] >= start_dt) & 
                    (data['income']['Date'] <= end_dt)
                ]
            
            if not data['expenses'].empty and 'Date' in data['expenses'].columns:
                data['expenses'] = data['expenses'][
                    (data['expenses']['Date'] >= start_dt) & 
                    (data['expenses']['Date'] <= end_dt)
                ]
        
        # Calculate totals
        total_income = data['income']['Amount'].sum() if not data['income'].empty else 0
        total_expenses = data['expenses']['Amount'].sum() if not data['expenses'].empty else 0
        net_cash_flow = total_income - total_expenses
        
        # Category breakdowns
        income_by_category = data['income'].groupby('Category')['Amount'].sum().to_dict() if not data['income'].empty else {}
        expenses_by_category = data['expenses'].groupby('Category')['Amount'].sum().to_dict() if not data['expenses'].empty else {}
        
        # Planned expenses
        planned_total = data['planned']['Estimated Cost'].sum() if not data['planned'].empty else 0
        
        return {
            'total_income': total_income,
            'total_expenses': total_expenses,
            'net_cash_flow': net_cash_flow,
            'income_by_category': income_by_category,
            'expenses_by_category': expenses_by_category,
            'planned_expenses': planned_total,
            'projected_cash_flow': net_cash_flow - planned_total
        }
    
    def get_monthly_trends(self):
        """Get monthly financial trends"""
        data = self.load_data()
        
        trends = []
        
        # Process income trends
        if not data['income'].empty and 'Date' in data['income'].columns:
            income_monthly = data['income'].groupby(data['income']['Date'].dt.to_period('M')).agg({
                'Amount': 'sum'
            }).reset_index()
            income_monthly['Type'] = 'Income'
            income_monthly['Month'] = income_monthly['Date'].astype(str)
            trends.extend(income_monthly[['Month', 'Amount', 'Type']].to_dict('records'))
        
        # Process expense trends
        if not data['expenses'].empty and 'Date' in data['expenses'].columns:
            expense_monthly = data['expenses'].groupby(data['expenses']['Date'].dt.to_period('M')).agg({
                'Amount': 'sum'
            }).reset_index()
            expense_monthly['Type'] = 'Expenses'
            expense_monthly['Month'] = expense_monthly['Date'].astype(str)
            expense_monthly['Amount'] = -expense_monthly['Amount']  # Negative for expenses
            trends.extend(expense_monthly[['Month', 'Amount', 'Type']].to_dict('records'))
        
        return trends