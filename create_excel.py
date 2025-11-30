from openpyxl import Workbook

# Create workbook
wb = Workbook()

# Income sheet
ws_income = wb.active
ws_income.title = "Income"
income_data = [
    ["Income Source", "Amount", "Date", "Category", "Notes"],
    ["Gumroad", 2070, "2025-11-03", "Digital Products", "Course sales"],
    ["Gumroad", 2080, "2025-11-10", "Digital Products", "Template sales"],
    ["Play Console", 9060, "2025-11-17", "App Revenue", "Monthly payout"],
    ["Runnable", 2925, "2025-11-18", "Services", "Development work"],
    ["Runnable", 2948, "2025-11-26", "Services", "Consulting"]
]
for row in income_data:
    ws_income.append(row)

# Expenses sheet
ws_expenses = wb.create_sheet("Expenses")
expense_data = [
    ["Description", "Amount", "Date", "Category", "Type", "Notes"],
    ["Cursor Pro", 2088.64, "2025-11-07", "Tools", "Software", "Annual subscription"],
    ["Editor Payment", 500, "2025-11-11", "Outsourcing", "Service", "Content editing"],
    ["Editor Payment", 500, "2025-11-25", "Outsourcing", "Service", "Content editing"],
    ["Basharat Salary", 4000, "2025-11-30", "Salaries", "Employee", "Monthly salary"],
    ["Abdaal Salary", 4000, "2025-11-30", "Salaries", "Employee", "Monthly salary"],
    ["Editor Payment", 500, "2025-11-30", "Outsourcing", "Service", "Content editing"],
    ["Saleem Salary", 4000, "2025-11-30", "Salaries", "Employee", "Monthly salary"]
]
for row in expense_data:
    ws_expenses.append(row)

# Planned sheet
ws_planned = wb.create_sheet("Planned")
planned_data = [
    ["Activity", "Estimated Cost", "Priority", "Status", "Target Date", "Notes"],
    ["Zoho Workplace (5 accounts)", 750, "High", "Pending", "2025-12-15", "Email & collaboration"],
    ["Domain Renewal", 1000, "High", "Pending", "2025-12-31", "Annual renewal"],
    ["Cursor Pro Renewal", 2000, "Medium", "Pending", "2026-11-07", "Development tool"],
    ["Editors Monthly Salary", 5000, "High", "Recurring", "2025-12-01", "Content team"],
    ["OpenAI API Credits", 750, "Medium", "Pending", "2025-12-10", "AI services"],
    ["Marketing Ads", 1000, "Low", "Planned", "2025-12-20", "Promotion campaign"]
]
for row in planned_data:
    ws_planned.append(row)

wb.save("Aivellum_Financials_OCT.xlsx")
print("Excel file created successfully!")